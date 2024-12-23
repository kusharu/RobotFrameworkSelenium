import json
import openpyxl   # pip install openpyxl
from pathlib import Path
import jsonpath  # pip install jsonpath

# get the Project path
val = Path(__file__).parent.parent
print(val)
print(type(val))
print(str(val))


def read_locator_from_json(locatorname):
    # locatorname is the KEY in ElementLocator.py whose value will be picked by this function
    # First open the JSON file

    f = open(str(val)+"\\JSONFiles\\ElementsRediffmail.json")
    # The read the file with read() method
    response = json.loads(f.read())
    value = jsonpath.jsonpath(response, locatorname)
    # When we read data using jsonpath it returns a list of one item ot more than that. Will return value at index 0
    return value[0]

valKEY = read_locator_from_json("Inbox.writemail_xpath");
print(valKEY)

print("***************************************************")

# print(str(val)+'\\TestCases\\DataDriven\\TestDataDriven.xlsx')
workBookPath = str(val)+'\\ExcelFile\\TestDataDriven.xlsx'
wb = openpyxl.load_workbook(workBookPath)


def get_max_row_number(sheetName):
    sh = wb[sheetName]
    maxRow = sh.max_row
    return maxRow

def get_max_column_number(sheetName):
    sh = wb[sheetName]
    maxColumn = sh.max_column
    return maxColumn

def get_data_from_cell(sheetName, row , column):
    sh = wb[sheetName]
    # If we do not typecast row and column arguments to INTEGER, might show error
    data = sh.cell(int(row), int(column))
    return data.value


# Calling the get_max_column_number(sheetName) function
numRows = get_max_row_number('Data')
print(numRows) # 3


# Calling the get_max_column_number(sheetName)
numColumns = get_max_column_number('Data')
print(numColumns) # 5

# Calling the get_data_from_cell(sheetName, row , column)
data = get_data_from_cell('Data', 2,1)
print(data) #selenium.testmay2017

# Calling the get_data_from_cell(sheetName, row , column)
data = get_data_from_cell('Data', 3,4)
print(data) # Development



